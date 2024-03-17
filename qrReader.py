import cv2
import qrcode
from tkinter import messagebox  
from openpyxl import load_workbook
import datetime
ExcelFilePath = "Filepath.xlsx"
wb = load_workbook(ExcelFilePath)
sheet = wb.worksheets[0]
columns = sheet["B"]

# import the opencv library
# define a video capture object
vid = cv2.VideoCapture(0)
detector = cv2.QRCodeDetector()
j = 3
while j < 15:
    if(sheet.cell(1, j).value.strftime("%x") == datetime.datetime.now().strftime("%x")):
        break
    j += 1

while True:
    # Capture the video frame by frame
    ret, frame = vid.read()
    data, bbox, straight_qrcode = detector.detectAndDecode(frame)
    i = 1
    if len(data) > 0:
        for d in columns:
            if(str(d.value) in str(data)):
                
                #get date value
                sheet.cell(i, j).value = 1
                wb.save(ExcelFilePath)
                break
            i += 1
        messagebox.showinfo("Info", data)
        print(data)

    # Display the resulting frame
    cv2.imshow('QrAttendance', frame)
    # the 'q' button is set as the
    # quitting button you may use any
    # desired button of your choice
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break
# After the loop release the cap object
vid.release()
# Destroy all the windows
cv2.destroyAllWindows()
